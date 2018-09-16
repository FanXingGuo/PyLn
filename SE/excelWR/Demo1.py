from openpyxl import load_workbook

wb=load_workbook("成绩.xlsx")
sheet=wb.active
A1_tab=sheet['A1']
a1=sheet.cell(row=1,column=1)

sheet['I2']='1'

wb.save("成绩.xlsx")
